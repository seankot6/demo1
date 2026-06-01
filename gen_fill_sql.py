# -*- coding: utf-8 -*-
import os
from datetime import datetime

import openpyxl

IMPORT_DIR = r"C:\Users\Пользователь\Desktop\redexpert\_pril_v1\pril2_data\import"
OUT = r"C:\BookStore\module1_fill_data.sql"


def esc(x):
    if x is None:
        return "NULL"
    s = str(x).strip()
    return "'" + s.replace("'", "''") + "'"


def parse_date(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if s == "30.02.2024":
        return "2024-02-29"
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    # fallback (чтобы не падать)
    return "2024-01-01"


def parse_items(s):
    parts = [p.strip() for p in str(s).split(",")]
    pairs = []
    i = 0
    while i < len(parts) - 1:
        pairs.append((parts[i], int(parts[i + 1])))
        i += 2
    return pairs


def main():
    # --- Tovar.xlsx ---
    wb = openpyxl.load_workbook(os.path.join(IMPORT_DIR, "Tovar.xlsx"), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    cats = []
    mans = []
    sups = []
    cat_set = set()
    man_set = set()
    sup_set = set()

    for r in rows:
        if not r[0]:
            continue
        cat = r[6]
        man = r[5]
        sup = r[4]
        if cat not in cat_set:
            cat_set.add(cat)
            cats.append(cat)
        if man not in man_set:
            man_set.add(man)
            mans.append(man)
        if sup not in sup_set:
            sup_set.add(sup)
            sups.append(sup)

    cat_id = {name: i + 1 for i, name in enumerate(cats)}
    man_id = {name: i + 1 for i, name in enumerate(mans)}
    sup_id = {name: i + 1 for i, name in enumerate(sups)}

    products = []
    product_id = 1
    for r in rows:
        if not r[0]:
            continue
        art, name, unit, price, sup, man, cat, disc, qty, desc, photo = r
        products.append(
            (
                product_id,
                art,
                name,
                unit,
                price,
                cat,
                man,
                sup,
                disc,
                qty,
                desc,
                photo,
            )
        )
        product_id += 1

    # --- Пункты выдачи_import.xlsx ---
    wb = openpyxl.load_workbook(
        os.path.join(IMPORT_DIR, "Пункты выдачи_import.xlsx"), read_only=True
    )
    ws = wb.active
    pp_rows = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    wb.close()

    # --- user_import.xlsx ---
    wb = openpyxl.load_workbook(os.path.join(IMPORT_DIR, "user_import.xlsx"), read_only=True)
    ws = wb.active
    user_rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    role_map = {
        "Администратор": 1,
        "Менеджер": 2,
        "Авторизированный клиент": 3,
    }
    status_map = {
        "Завершен": 1,
        "Новый": 2,
    }

    lines = []
    lines.append("-- Auto-fill for Module 1: INSERT data into tables")
    lines.append("-- Run on empty DB (unique constraints will conflict otherwise).")
    lines.append("")

    # ROLE
    lines.append("INSERT INTO ROLE (ROLE_ID, ROLE_NAME) VALUES (1, 'Администратор');")
    lines.append("INSERT INTO ROLE (ROLE_ID, ROLE_NAME) VALUES (2, 'Менеджер');")
    lines.append("INSERT INTO ROLE (ROLE_ID, ROLE_NAME) VALUES (3, 'Авторизированный клиент');")
    lines.append("")

    # ORDER_STATUS
    lines.append("INSERT INTO ORDER_STATUS (STATUS_ID, STATUS_NAME) VALUES (1, 'Завершен');")
    lines.append("INSERT INTO ORDER_STATUS (STATUS_ID, STATUS_NAME) VALUES (2, 'Новый');")
    lines.append("")

    # CATEGORY
    for name, i in cat_id.items():
        lines.append(
            f"INSERT INTO CATEGORY (CATEGORY_ID, CATEGORY_NAME) VALUES ({i}, {esc(name)});"
        )
    lines.append("")

    # MANUFACTURER
    for name, i in man_id.items():
        lines.append(
            f"INSERT INTO MANUFACTURER (MANUFACTURER_ID, MANUFACTURER_NAME) VALUES ({i}, {esc(name)});"
        )
    lines.append("")

    # SUPPLIER
    for name, i in sup_id.items():
        lines.append(
            f"INSERT INTO SUPPLIER (SUPPLIER_ID, SUPPLIER_NAME) VALUES ({i}, {esc(name)});"
        )
    lines.append("")

    # PICKUP_POINT
    for i, addr in enumerate(pp_rows, start=1):
        lines.append(
            f"INSERT INTO PICKUP_POINT (PICKUP_POINT_ID, ADDRESS) VALUES ({i}, {esc(addr)});"
        )
    lines.append("")

    # USERS (USER_ID = row index 1..)
    for i, (role, fio, login, pwd) in enumerate(user_rows, start=1):
        rid = role_map.get(str(role).strip())
        if rid is None:
            raise RuntimeError(f"Unknown role in user_import.xlsx: {role}")
        lines.append(
            "INSERT INTO USERS (USER_ID, ROLE_ID, FULL_NAME, LOGIN, PASSWORD) VALUES (%d, %d, %s, %s, %s);"
            % (i, rid, esc(fio), esc(login), esc(pwd))
        )
    lines.append("")

    # PRODUCT (PRODUCT_ID = row index 1..)
    for (
        pid,
        art,
        name,
        unit,
        price,
        cat,
        man,
        sup,
        disc,
        qty,
        desc,
        photo,
    ) in products:
        photo_path = f"import\\{photo}" if photo else None
        disc_v = 0 if disc is None else disc
        qty_v = 0 if qty is None else int(qty)
        lines.append(
            "INSERT INTO PRODUCT (PRODUCT_ID, ARTICLE, PRODUCT_NAME, UNIT_MEASURE, PRICE, CATEGORY_ID, MANUFACTURER_ID, SUPPLIER_ID, DISCOUNT, QUANTITY_STOCK, DESCRIPTION, PHOTO_PATH) VALUES (%d, %s, %s, %s, %s, %d, %d, %d, %s, %d, %s, %s);"
            % (
                pid,
                esc(art),
                esc(name),
                esc(unit),
                str(price).replace(",", "."),
                cat_id[cat],
                man_id[man],
                sup_id[sup],
                str(disc_v).replace(",", "."),
                qty_v,
                esc(desc),
                esc(photo_path),
            )
        )
    lines.append("")

    # ORDERS + ORDER_ITEM
    wb = openpyxl.load_workbook(os.path.join(IMPORT_DIR, "Заказ_import.xlsx"), read_only=True)
    ws = wb.active
    order_rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    order_item_id = 1
    # helper maps for FK by ARTICLE / FULL_NAME
    art_to_pid = {str(p[1]).strip(): p[0] for p in products}
    fio_to_uid = {str(r[1]).strip(): idx for idx, r in enumerate(user_rows, start=1)}

    for r in order_rows:
        if not r[0]:
            continue
        oid, items, od, dd, pp, fio, code, status = r[:8]
        st = str(status).strip() if status is not None else "Новый"
        st_norm = "Новый" if st.startswith("Новый") else st
        sid = status_map.get(st_norm, 2)
        uid = fio_to_uid.get(str(fio).strip(), 1)

        lines.append(
            "INSERT INTO ORDERS (ORDER_ID, ORDER_DATE, DELIVERY_DATE, PICKUP_POINT_ID, USER_ID, PICKUP_CODE, STATUS_ID) VALUES (%d, %s, %s, %d, %d, %d, %d);"
            % (
                int(oid),
                esc(parse_date(od)),
                esc(parse_date(dd)) if dd is not None else "NULL",
                int(pp),
                uid,
                int(code),
                sid,
            )
        )

        for art, qty in parse_items(items):
            pid = art_to_pid.get(str(art).strip())
            if not pid:
                continue
            lines.append(
                "INSERT INTO ORDER_ITEM (ORDER_ITEM_ID, ORDER_ID, PRODUCT_ID, QUANTITY) VALUES (%d, %d, %d, %d);"
                % (order_item_id, int(oid), pid, int(qty))
            )
            order_item_id += 1

    with open(OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print("Wrote:", OUT, "lines:", len(lines))


if __name__ == "__main__":
    main()

